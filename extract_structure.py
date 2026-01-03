from openpyxl import load_workbook
import json

# 读取Excel文件
wb = load_workbook('kaoshibaoExcel20221101.xlsx')
ws = wb.active

# 提取表头
headers = []
for cell in ws[1]:
    if cell.value:
        headers.append(cell.value)
    else:
        break

print("表头列数:", len(headers))
print("\n表头信息:")
for idx, header in enumerate(headers, 1):
    print(f"{idx}. {header}")

# 提取所有数据行
print("\n\n数据行分析:")
all_data = []
for row_idx in range(2, ws.max_row + 1):
    row_data = {}
    for col_idx, header in enumerate(headers, 1):
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        row_data[header] = cell_value
    all_data.append(row_data)

# 保存为JSON文件以便查看
with open('extracted_data.json', 'w', encoding='utf-8') as f:
    json.dump({
        'headers': headers,
        'data': all_data
    }, f, ensure_ascii=False, indent=2)

print(f"共提取 {len(all_data)} 行数据")
print("数据已保存到 extracted_data.json")
