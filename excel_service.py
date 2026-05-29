#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json
import os
import tempfile
import zipfile
from io import BytesIO
from copy import copy
from xml.sax.saxutils import escape
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from generate_excel import create_excel_from_json
from header_utils import match_header


TEMPLATE_STANDARD = "standard"
TEMPLATE_ANSWER_HELPER = "answer_helper"
TEMPLATE_DIR = "templates"
ANSWER_HELPER_TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "题目导入模板(答题帮手).xlsx")
WORD_EXPORT_TEMPLATE_PATH = os.path.join(TEMPLATE_DIR, "Word导出模板_考试答案解析版.docx")
ANSWER_HELPER_HEADERS = [
    "题型", "题干内容", "题目解析", "题目难度", "章节", "真题", "知识点",
    "正确答案", "选项A", "选项B", "选项C", "选项D", "选项E"
]


def normalize_template_name(template):
    if template in (TEMPLATE_ANSWER_HELPER, "答题帮手", "answer-helper"):
        return TEMPLATE_ANSWER_HELPER
    return TEMPLATE_STANDARD


def detect_question_template(question):
    keys = set(question.keys())
    answer_helper_hits = len(keys & set(ANSWER_HELPER_HEADERS))
    standard_hits = len([
        key for key in keys
        if any(marker in key for marker in ("题干（必填）", "题型 （必填）", "选项 A", "正确答案\n（必填）"))
    ])
    if answer_helper_hits > standard_hits:
        return TEMPLATE_ANSWER_HELPER
    return TEMPLATE_STANDARD


def question_value(question, header):
    return match_header(header, question)


def normalize_difficulty(value):
    mapping = {
        "易": "容易",
        "偏易": "较易",
        "适中": "中等",
        "中": "中等",
        "难": "较难",
        "困难": "困难",
    }
    text = str(value or "").strip()
    return mapping.get(text, text or "中等")


def normalize_answer_for_answer_helper(question_type, answer):
    text = str(answer or "").strip()
    if question_type == "判断题":
        if text in ("A", "对", "正确", "True", "true"):
            return "正确"
        if text in ("B", "错", "错误", "False", "false"):
            return "错误"
    return text


def normalize_answer_for_standard(question_type, answer):
    text = str(answer or "").strip()
    if question_type == "判断题":
        if text in ("正确", "对", "True", "true"):
            return "A"
        if text in ("错误", "错", "False", "false"):
            return "B"
    return text


def standard_question_to_answer_helper(question):
    question_type = question_value(question, "题型 （必填）")
    if detect_question_template(question) == TEMPLATE_ANSWER_HELPER:
        return {
            "题型": question.get("题型"),
            "题干内容": question.get("题干内容"),
            "题目解析": question.get("题目解析"),
            "题目难度": normalize_difficulty(question.get("题目难度")),
            "章节": question.get("章节"),
            "真题": question.get("真题"),
            "知识点": question.get("知识点"),
            "正确答案": normalize_answer_for_answer_helper(question.get("题型"), question.get("正确答案")),
            "选项A": question.get("选项A"),
            "选项B": question.get("选项B"),
            "选项C": question.get("选项C"),
            "选项D": question.get("选项D"),
            "选项E": question.get("选项E"),
        }
    return {
        "题型": question_type,
        "题干内容": question_value(question, "题干（必填）"),
        "题目解析": question_value(question, "解析\n（勿删）"),
        "题目难度": normalize_difficulty(question_value(question, "难度")),
        "章节": question_value(question, "章节\n（勿删）"),
        "真题": question.get("真题"),
        "知识点": question.get("知识点"),
        "正确答案": normalize_answer_for_answer_helper(question_type, question_value(question, "正确答案\n（必填）")),
        "选项A": question_value(question, "选项 A"),
        "选项B": question_value(question, "选项 B"),
        "选项C": question_value(question, "选项 C"),
        "选项D": question_value(question, "选项 D"),
        "选项E": question_value(question, "选项E\n(勿删)"),
    }


def answer_helper_question_to_standard(question):
    if detect_question_template(question) == TEMPLATE_STANDARD:
        question_type = question_value(question, "题型 （必填）")
        return {
            "题干（必填）": question_value(question, "题干（必填）"),
            "题型 （必填）": question_type,
            "选项 A": question_value(question, "选项 A"),
            "选项 B": question_value(question, "选项 B"),
            "选项 C": question_value(question, "选项 C"),
            "选项 D": question_value(question, "选项 D"),
            "选项E\n(勿删)": question_value(question, "选项E\n(勿删)"),
            "选项F\n(勿删)": question_value(question, "选项F\n(勿删)"),
            "选项G\n(勿删)": question_value(question, "选项G\n(勿删)"),
            "选项H\n(勿删)": question_value(question, "选项H\n(勿删)"),
            "正确答案\n（必填）": normalize_answer_for_standard(question_type, question_value(question, "正确答案\n（必填）")),
            "解析\n（勿删）": question_value(question, "解析\n（勿删）"),
            "章节\n（勿删）": question_value(question, "章节\n（勿删）"),
            "难度": question_value(question, "难度"),
        }
    question_type = question.get("题型")
    return {
        "题干（必填）": question.get("题干内容"),
        "题型 （必填）": question_type,
        "选项 A": question.get("选项A"),
        "选项 B": question.get("选项B"),
        "选项 C": question.get("选项C"),
        "选项 D": question.get("选项D"),
        "选项E\n(勿删)": question.get("选项E"),
        "选项F\n(勿删)": question.get("选项F"),
        "选项G\n(勿删)": question.get("选项G"),
        "选项H\n(勿删)": question.get("选项H"),
        "正确答案\n（必填）": normalize_answer_for_standard(question_type, question.get("正确答案")),
        "解析\n（勿删）": question.get("题目解析"),
        "章节\n（勿删）": question.get("章节"),
        "难度": question.get("题目难度"),
    }


def convert_questions(questions, target_template):
    target = normalize_template_name(target_template)
    if target == TEMPLATE_ANSWER_HELPER:
        return [standard_question_to_answer_helper(question) for question in questions]
    return [answer_helper_question_to_standard(question) for question in questions]


def copy_row_style(source_row, target_row):
    for source_cell, target_cell in zip(source_row, target_row):
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)


def parse_excel_to_questions(file_bytes):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    row1_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    row2_headers = [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    row2_hits = len(set(row2_headers) & set(ANSWER_HELPER_HEADERS))
    row1_hits = len([header for header in row1_headers if header])

    if row2_hits >= 3:
        source_template = TEMPLATE_ANSWER_HELPER
        header_row = 2
        start_row = 3
        headers = row2_headers
    else:
        source_template = TEMPLATE_STANDARD
        header_row = 1
        start_row = 2
        headers = row1_headers

    questions = []
    for row_idx in range(start_row, ws.max_row + 1):
        question = {}
        has_value = False
        for col_idx, header in enumerate(headers, 1):
            if not header:
                continue
            value = ws.cell(row_idx, col_idx).value
            if value is not None and str(value).strip() != "":
                has_value = True
            question[str(header)] = value
        if has_value:
            questions.append(question)

    return {
        "sourceTemplate": source_template,
        "questions": convert_questions(questions, TEMPLATE_STANDARD)
    }


def create_answer_helper_excel(questions, output_file):
    if os.path.exists(ANSWER_HELPER_TEMPLATE_PATH):
        wb = load_workbook(ANSWER_HELPER_TEMPLATE_PATH)
        ws = wb.active
        start_row = 3
        template_row = 3 if ws.max_row >= 3 else 2
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None
    else:
        wb, ws = create_answer_helper_workbook()
        start_row = 3
        template_row = 2

    header_row = 2
    headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
    if not any(headers):
        headers = ANSWER_HELPER_HEADERS
        for col, header in enumerate(headers, 1):
            ws.cell(header_row, col).value = header

    converted = [standard_question_to_answer_helper(question) for question in questions]
    for offset, question in enumerate(converted):
        row_idx = start_row + offset
        if template_row and row_idx != template_row:
            copy_row_style(ws[template_row], ws[row_idx])
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx).value = question.get(header)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(ws.column_dimensions[get_column_letter(col_idx)].width or 12, 12)

    wb.save(output_file)


def create_answer_helper_workbook():
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"
    ws.cell(row=1, column=1).value = "答题帮手导入模板"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ANSWER_HELPER_HEADERS))
    ws.row_dimensions[1].height = 42
    header_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, header in enumerate(ANSWER_HELPER_HEADERS, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(name="宋体", size=11, bold=True)
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return wb, ws


def export_to_excel(questions, template=TEMPLATE_STANDARD):
    """导出题目到 Excel 文件"""
    template = normalize_template_name(template)
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
        json.dump({'questions': questions}, f, ensure_ascii=False, indent=2)
        json_path = f.name

    excel_path = tempfile.mktemp(suffix='.xlsx')

    try:
        if template == TEMPLATE_ANSWER_HELPER:
            create_answer_helper_excel(questions, excel_path)
        else:
            questions = [answer_helper_question_to_standard(question) for question in questions]
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump({'questions': questions}, f, ensure_ascii=False, indent=2)
            create_excel_from_json(json_path, excel_path)
        return excel_path, json_path
    except Exception as e:
        if os.path.exists(json_path):
            os.unlink(json_path)
        raise e


def _docx_paragraph(text, style=None):
    text = "" if text is None else str(text)
    lines = text.splitlines() or [""]
    escaped_lines = [escape(line) for line in lines]
    runs = "<w:br/>".join(f"<w:t xml:space=\"preserve\">{line}</w:t>" for line in escaped_lines)
    style_xml = f"<w:pPr><w:pStyle w:val=\"{style}\"/></w:pPr>" if style else ""
    return f"<w:p>{style_xml}<w:r>{runs}</w:r></w:p>"


def _docx_blank_lines(count):
    return "".join("<w:p/>" for _ in range(count))


def export_to_word(questions):
    normalized_questions = [answer_helper_question_to_standard(question) for question in questions]
    body_parts = [
        _docx_paragraph("题库导出文档", "Title"),
        _docx_paragraph("考试卷面 + 答案解析版", "Subtitle"),
        _docx_paragraph(f"题目数量：{len(normalized_questions)}"),
    ]
    answer_summary = []

    previous_chapter = None
    for idx, question in enumerate(normalized_questions, 1):
        question_type = question_value(question, "题型 （必填）") or "题目"
        stem = question_value(question, "题干（必填）") or ""
        answer = question_value(question, "正确答案\n（必填）") or ""
        analysis = question_value(question, "解析\n（勿删）") or ""
        chapter = question_value(question, "章节\n（勿删）") or ""
        difficulty = question_value(question, "难度") or ""

        if chapter and chapter != previous_chapter:
            if previous_chapter is not None:
                body_parts.append(_docx_blank_lines(3))
            body_parts.append(_docx_paragraph(chapter, "Section"))
            previous_chapter = chapter
        elif idx > 1:
            body_parts.append(_docx_blank_lines(2))

        body_parts.append(_docx_paragraph(f"{idx}. 【{question_type}】{stem}", "Question"))
        for option_header in ("选项 A", "选项 B", "选项 C", "选项 D", "选项E\n(勿删)", "选项F\n(勿删)", "选项G\n(勿删)", "选项H\n(勿删)"):
            value = question_value(question, option_header)
            if value is not None and str(value).strip():
                option_label = option_header.replace("选项 ", "").replace("选项", "").replace("\n(勿删)", "").strip()
                body_parts.append(_docx_paragraph(f"{option_label}. {value}"))
        body_parts.append(_docx_paragraph(f"【答案】{answer}", "Answer"))
        if analysis:
            body_parts.append(_docx_paragraph(f"【解析】{analysis}", "Analysis"))
        meta_parts = [item for item in (f"章节：{chapter}" if chapter else "", f"难度：{difficulty}" if difficulty else "") if item]
        if meta_parts:
            body_parts.append(_docx_paragraph("；".join(meta_parts), "Analysis"))
        answer_summary.append(_docx_paragraph(f"{idx}. {answer}"))

    if answer_summary:
        body_parts.append('<w:p><w:r><w:br w:type="page"/></w:r></w:p>')
        body_parts.append(_docx_paragraph("答案汇总", "Section"))
        body_parts.extend(answer_summary)

    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    {''.join(body_parts)}
    <w:sectPr><w:pgSz w:w="11906" w:h="16838"/><w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/></w:sectPr>
  </w:body>
</w:document>"""

    output_path = tempfile.mktemp(suffix=".docx")
    if os.path.exists(WORD_EXPORT_TEMPLATE_PATH):
        with zipfile.ZipFile(WORD_EXPORT_TEMPLATE_PATH, "r") as template_docx:
            with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as output_docx:
                for item in template_docx.infolist():
                    if item.filename == "word/document.xml":
                        continue
                    output_docx.writestr(item, template_docx.read(item.filename))
                output_docx.writestr("word/document.xml", document_xml)
    else:
        content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
        rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
        with zipfile.ZipFile(output_path, "w", zipfile.ZIP_DEFLATED) as docx:
            docx.writestr("[Content_Types].xml", content_types)
            docx.writestr("_rels/.rels", rels)
            docx.writestr("word/document.xml", document_xml)
    return output_path
