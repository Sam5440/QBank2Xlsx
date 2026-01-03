from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
import json

# 读取Excel文件
wb = load_workbook('kaoshibaoExcel20221101.xlsx')
ws = wb.active

print(f"工作表名称: {ws.title}")
print(f"最大行数: {ws.max_row}")
print(f"最大列数: {ws.max_column}")
print("\n" + "="*80)

# 读取前15行数据
print("\n前15行数据:")
for i in range(1, min(16, ws.max_row + 1)):
    row_data = []
    for cell in ws[i]:
        row_data.append(cell.value)
    print(f"第{i}行: {row_data[:14]}")  # 只显示前14列

print("\n" + "="*80)

# 分析表格结构
print("\n表格结构分析:")
print("第1行（表头）:")
headers = [cell.value for cell in ws[1]]
for idx, header in enumerate(headers[:14], 1):
    print(f"  列{idx}: {header}")
