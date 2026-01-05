#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
根据JSON数据生成考试题库Excel文件
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 表头映射表（统一管理所有字段）
HEADER_MAPPING = {
    '题干': "题干（必填）",
    '题型': "题型 （必填）",
    'A': "选项 A",
    'B': "选项 B",
    'C': "选项 C",
    'D': "选项 D",
    'E': "选项E\n(勿删)",
    'F': "选项F\n(勿删)",
    'G': "选项G\n(勿删)",
    'H': "选项H\n(勿删)",
    '正确答案': "正确答案\n（必填）",
    '解析': "解析\n（勿删）",
    '章节': "章节\n（勿删）",
    '难度': "难度"
}

# 固定表头（永远不变）
HEADERS = list(HEADER_MAPPING.values())


def match_header(header, question):
    """
    灵活匹配表头，尝试多种可能的键格式

    参数:
        header: 标准表头字符串
        question: 题目字典

    返回:
        匹配到的值，如果不存在返回 None
    """
    # 提取括号前的核心文本
    base_text = header.split('（')[0].split('(')[0].replace('\n', '').strip()

    # 尝试多种可能的键格式
    possible_keys = [
        header,                           # 完整格式（如 "正确答案\n（必填）"）
        header.replace('\n', ''),         # 去除换行符（如 "正确答案（必填）"）
        base_text                         # 只有核心文本（如 "正确答案"）
    ]

    for key in possible_keys:
        value = question.get(key)
        if value is not None:
            return value

    return None


def load_json_data(json_file):
    """加载JSON数据"""
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def create_excel_from_json(json_file, output_file):
    """
    从JSON文件创建Excel文件

    参数:
        json_file: JSON数据文件路径
        output_file: 输出的Excel文件路径
    """
    # 加载数据
    data = load_json_data(json_file)
    questions = data.get('questions', [])

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "题库（答案请直接导入）"

    # 写入表头并设置样式
    setup_header(ws, HEADERS)

    # 写入数据
    write_data(ws, questions, HEADERS)

    # 调整列宽
    adjust_column_width(ws)

    # 保存文件
    wb.save(output_file)
    print(f"Excel文件已生成: {output_file}")
    print(f"共导入 {len(questions)} 道题目")


def setup_header(ws, headers):
    """设置表头样式"""
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 定义表头填充色（浅蓝色）
    header_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")

    # 写入表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(name='宋体', size=11, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = header_fill

    # 设置表头行高
    ws.row_dimensions[1].height = 40


def write_data(ws, questions, headers):
    """写入题目数据"""
    # 定义边框样式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 从第2行开始写入数据
    for row_idx, question in enumerate(questions, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = match_header(header, question)
            cell.font = Font(name='宋体', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell.border = thin_border

        # 设置数据行高
        ws.row_dimensions[row_idx].height = 30


def adjust_column_width(ws):
    """调整列宽"""
    # 定义各列的宽度
    column_widths = {
        1: 50,   # 题干
        2: 15,   # 题型
        3: 20,   # 选项A
        4: 20,   # 选项B
        5: 20,   # 选项C
        6: 20,   # 选项D
        7: 20,   # 选项E
        8: 20,   # 选项F
        9: 20,   # 选项G
        10: 20,  # 选项H
        11: 15,  # 正确答案
        12: 30,  # 解析
        13: 20,  # 章节
        14: 10   # 难度
    }

    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


if __name__ == "__main__":
    # 使用示例
    json_file = "demo_questions.json"  # 输入的JSON文件
    output_file = "generated_exam.xlsx"  # 输出的Excel文件

    try:
        create_excel_from_json(json_file, output_file)
    except FileNotFoundError:
        print(f"错误: 找不到文件 {json_file}")
    except Exception as e:
        print(f"错误: {e}")
